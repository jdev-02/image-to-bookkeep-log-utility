"""XLSX writer with yellow highlighting and comments."""

from pathlib import Path
from typing import Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from itbl.output.writer_base import WriterBase


class XLSXWriter(WriterBase):
    """XLSX writer with highlighting support."""

    def __init__(self, highlight_color: str = "#FFF59D"):
        """
        Initialize XLSX writer.
        
        Args:
            highlight_color: Hex color for highlights (default: yellow #FFF59D)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for XLSX output. Install: pip install openpyxl"
            )
        self.highlight_color = highlight_color
        self.fill = PatternFill(start_color=highlight_color.replace("#", ""), end_color=highlight_color.replace("#", ""), fill_type="solid")

    def write(
        self,
        rows: List[Dict],
        output_path: Path,
        category: str,
        apply_highlights: bool = False,
    ) -> None:
        """
        Write rows to XLSX file.
        
        Args:
            rows: List of normalized row dicts
            output_path: Output file path (or directory; will create category_name.xlsx)
            category: Category/tab name
            apply_highlights: Whether to apply yellow highlighting
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not available")

        if output_path.is_dir():
            output_file = output_path / f"{category.replace(' ', '_')}.xlsx"
        else:
            output_file = output_path

        if not rows:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = category

        # Determine columns (skip hidden fields)
        hidden_prefixes = ["_"]
        sample_row = rows[0]
        all_cols = set(sample_row.keys())
        visible_cols = [
            col for col in all_cols
            if not any(col.startswith(prefix) for prefix in hidden_prefixes)
        ]
        visible_cols = sorted(list(visible_cols))

        # Write header
        for col_idx, col_name in enumerate(visible_cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = cell.font.copy(bold=True)

        # Write rows
        for row_idx, row in enumerate(rows, start=2):
            flags = row.get("_flags", [])
            highlight_cells = row.get("_highlight_cells", [])

            for col_idx, col_name in enumerate(visible_cols, start=1):
                value = row.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply highlighting if flagged
                if apply_highlights and col_name in highlight_cells:
                    cell.fill = self.fill
                    # Add comment with reason
                    if flags:
                        comment_text = f"Reason: {', '.join(flags)}\nConfidence: {row.get('_ocr_confidence', 'N/A')}"
                        if row.get("_low_conf_tokens"):
                            comment_text += f"\nLow confidence tokens detected"
                        cell.comment = Comment(comment_text, "itbl")

        wb.save(output_file)

